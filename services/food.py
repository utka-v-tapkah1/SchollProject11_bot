import logging

from services.playwright_parses import parse_xlsx
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from lexicon.lexicon import LEXICON_RU
import os


download_path = "./xlsx_downloads"

logger = logging.getLogger(__name__)


async def _create_xlsx_downloads() -> None:
    if not os.path.isdir(download_path):
        os.mkdir(download_path)


async def _delete_files_in_folder(folder_path) -> None:
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f'Ошибка при удалении файла {file_path}. {e}')


async def _xlsx_file_enable():
    if not os.listdir(download_path):
        return False
    return True


async def get_xlsx():
    # Создание папки для xlsx файлов, если таковая отсутствует
    await _create_xlsx_downloads()

    today = datetime.now().strftime('%Y-%m-%d')
    # Если уже есть файл, то ничего не парсим
    for file in os.listdir(download_path):
        if file == f'{today}-sm.xlsx':
            return True
    await _delete_files_in_folder(download_path)
    await parse_xlsx(today)

    return await _xlsx_file_enable()


async def get_food_breakfast():
    flag = await get_xlsx()
    if not flag:
        return LEXICON_RU['no_breakfast']
    path = f'{download_path}/{os.listdir(download_path)[0]}'
    workbook = openpyxl.load_workbook(path)
    sheet: Worksheet = workbook.active

    barrier = '-' * 36
    a, b, c = "Раздел",	"№ рец.", "Блюдо"
    rows = (f'Завтрак\n'
            f'{barrier}\n'
            f'{a} | <i>{b}</i> | <b>{c}</b>\n'
            f'{barrier}\n')

    for row in sheet.iter_rows(min_row=4, max_row=10,
                               min_col=2, max_col=4,
                               values_only=True):
        if row[0] is None:
            continue
        a, b, c = row
        rows += f'{a} | <i>{b}</i> | <b>{c}</b>\n'
        rows += f'{barrier}\n'

    return rows


async def get_food_obed():
    flag = await get_xlsx()
    if not flag:
        return LEXICON_RU['no_obed']
    path = f'{download_path}/{os.listdir(download_path)[0]}'
    workbook = openpyxl.load_workbook(path)
    sheet: Worksheet = workbook.active

    barrier = '-' * 36
    a, b, c = "Раздел",	"№ рец.", "Блюдо"
    rows = (f'Обед\n'
            f'{barrier}\n'
            f'{a} | <i>{b}</i> | <b>{c}</b>\n'
            f'{barrier}\n')

    for row in sheet.iter_rows(min_row=14, max_row=22,
                               min_col=2, max_col=4,
                               values_only=True):
        if row[0] is None:
            continue
        a, b, c = row
        rows += f'{a} | <i>{b}</i> | <b>{c}</b>\n'
        rows += f'{barrier}\n'

    return rows
